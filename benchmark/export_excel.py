"""
benchmark/export_excel.py

Converts a benchmark CSV into a fully formatted Excel workbook with:
  Sheet 1 — Results Table       (formatted, colour-coded)
  Sheet 2 — Speed vs Accuracy   (scatter chart: FPS vs mAP@50)
  Sheet 3 — Accuracy Comparison (bar chart: mAP@50, Precision, Recall per model)
  Sheet 4 — Speed Comparison    (bar chart: FPS and Avg ms per model)
  Sheet 5 — Size vs Accuracy    (scatter: model size vs mAP@50)
  Sheet 6 — Raw Data            (unformatted copy for pivot tables)

Usage:
  python benchmark/export_excel.py --csv benchmark_results.csv
  python benchmark/export_excel.py --csv benchmark_results.csv --out results.xlsx
"""

import argparse
import csv
import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList


# ── Colour palette ───────────────────────────────────────────────────── #
C_HEADER_BG   = "1E1E2E"   # dark header background
C_HEADER_FG   = "CDD6F4"   # header text
C_YOLO_BG     = "1E3A5F"   # YOLO row accent
C_MOBILE_BG   = "1E3A2F"   # MobileNet row accent
C_ALT_BG      = "2A2A3E"   # alternating row
C_BEST_FPS    = "1A3D1A"   # best FPS highlight
C_BEST_MAP    = "1A2D4D"   # best mAP highlight
C_TEXT        = "D4D4D4"   # normal text
C_ACCENT      = "0078D4"   # blue accent
C_GREEN       = "4EC94E"
C_ORANGE      = "CE9178"
C_WHITE       = "FFFFFF"
C_DARK        = "1E1E1E"

YOLO_MODELS     = {"YOLO"}
MOBILENET_MODELS= {"MobileNet"}

# ── Speed-only columns vs eval columns ──────────────────────────────── #
SPEED_COLS = ["Model", "Family", "Size (MB)", "Avg ms", "FPS",
              "Avg dets", "Device", "N"]
EVAL_COLS  = ["Model", "Family", "Size (MB)", "Avg ms", "FPS",
              "mAP@50 (%)", "mAP@50:95 (%)", "Precision (%)", "Recall (%)",
              "Device", "N"]


def thin_border():
    s = Side(style="thin", color="3F3F46")
    return Border(left=s, right=s, top=s, bottom=s)


def header_font():
    return Font(name="Segoe UI", bold=True, color=C_HEADER_FG, size=10)


def cell_font(bold=False, color=C_TEXT):
    return Font(name="Segoe UI", bold=bold, color=color, size=10)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center")


# ── CSV reader ───────────────────────────────────────────────────────── #
def read_csv(path: str) -> tuple[list[str], list[dict]]:
    rows = []
    headers = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        for row in reader:
            # Skip rows where all numeric fields are empty (separator rows)
            numeric_fields = [k for k in headers if k not in ("Model", "Family", "Device")]
            if all(not row.get(k, "").strip() for k in numeric_fields):
                continue
            rows.append(row)
    return headers, rows


def is_eval_mode(headers: list[str]) -> bool:
    return "mAP@50 (%)" in headers


def safe_float(val: str) -> float | None:
    try:
        return float(val.strip()) if val and val.strip() not in ("", "—", "None") else None
    except ValueError:
        return None


# ── Sheet 1: Results Table ───────────────────────────────────────────── #
def build_results_sheet(ws, headers, rows, eval_mode):
    ws.title = "Results"
    ws.sheet_view.showGridLines = False
    ws.tab_color = C_ACCENT

    # Title
    ws.merge_cells("A1:K1" if eval_mode else "A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "ECP Pedestrian Detection — Model Benchmark Results"
    title_cell.font  = Font(name="Segoe UI", bold=True, color=C_WHITE, size=14)
    title_cell.fill  = fill(C_HEADER_BG)
    title_cell.alignment = center()
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:K2" if eval_mode else "A2:H2")
    sub = ws["A2"]
    sub.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  " \
                f"Dataset: ECP val  |  conf=0.35  IoU=0.50"
    sub.font  = Font(name="Segoe UI", color="9D9D9D", size=9, italic=True)
    sub.fill  = fill(C_HEADER_BG)
    sub.alignment = center()
    ws.row_dimensions[2].height = 18

    # Headers row
    col_widths = {
        "Model": 36, "Family": 12, "Size (MB)": 10, "Avg ms": 10,
        "FPS": 8, "mAP@50 (%)": 12, "mAP@50:95 (%)": 14,
        "Precision (%)": 13, "Recall (%)": 11, "Avg dets": 10,
        "Device": 10, "N": 6,
    }

    display_headers = EVAL_COLS if eval_mode else SPEED_COLS
    for c_idx, col_name in enumerate(display_headers, 1):
        cell = ws.cell(row=3, column=c_idx, value=col_name)
        cell.font      = header_font()
        cell.fill      = fill(C_HEADER_BG)
        cell.alignment = center()
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col_name, 12)
    ws.row_dimensions[3].height = 22

    # Find best values for highlighting
    fps_vals  = [safe_float(r.get("FPS", "")) for r in rows]
    map_vals  = [safe_float(r.get("mAP@50 (%)", "")) for r in rows] if eval_mode else []
    best_fps  = max((v for v in fps_vals  if v is not None), default=None)
    best_map  = max((v for v in map_vals  if v is not None), default=None)

    # Data rows
    for r_idx, row in enumerate(rows, 4):
        family   = row.get("Family", "").strip()
        fps_val  = safe_float(row.get("FPS", ""))
        map_val  = safe_float(row.get("mAP@50 (%)", "")) if eval_mode else None

        # Row background
        if map_val is not None and map_val == best_map:
            row_bg = C_BEST_MAP
        elif fps_val is not None and fps_val == best_fps:
            row_bg = C_BEST_FPS
        elif family == "YOLO":
            row_bg = C_YOLO_BG if r_idx % 2 == 0 else "1A2E50"
        elif family == "MobileNet":
            row_bg = C_MOBILE_BG if r_idx % 2 == 0 else "162E20"
        else:
            row_bg = C_ALT_BG if r_idx % 2 == 0 else C_DARK

        for c_idx, col_name in enumerate(display_headers, 1):
            val = row.get(col_name, "").strip()
            f   = safe_float(val)

            cell = ws.cell(row=r_idx, column=c_idx)
            cell.fill   = fill(row_bg)
            cell.border = thin_border()

            if col_name == "Model":
                cell.value     = val
                cell.font      = cell_font(bold=True, color=C_WHITE)
                cell.alignment = left()
            elif col_name in ("Family", "Device"):
                cell.value     = val
                cell.font      = cell_font(color="9CDCFE")
                cell.alignment = center()
            elif col_name == "N":
                cell.value     = int(f) if f else val
                cell.font      = cell_font(color="9D9D9D")
                cell.alignment = center()
            elif f is not None:
                cell.value     = f
                # colour-code metric cells
                if col_name == "FPS" and f == best_fps:
                    cell.font = cell_font(bold=True, color=C_GREEN)
                elif col_name == "mAP@50 (%)" and f == best_map:
                    cell.font = cell_font(bold=True, color="9CDCFE")
                elif col_name in ("mAP@50 (%)", "mAP@50:95 (%)",
                                  "Precision (%)", "Recall (%)"):
                    cell.font = cell_font(color="DCDCAA")
                elif col_name in ("Avg ms",):
                    cell.font = cell_font(color=C_ORANGE)
                else:
                    cell.font = cell_font(color=C_TEXT)
                cell.alignment = center()
                if col_name in ("mAP@50 (%)", "mAP@50:95 (%)",
                                "Precision (%)", "Recall (%)"):
                    cell.number_format = "0.0"
                elif col_name in ("Avg ms", "FPS"):
                    cell.number_format = "0.0"
                elif col_name == "Size (MB)":
                    cell.number_format = "0.0"
            else:
                cell.value     = val
                cell.font      = cell_font()
                cell.alignment = center()

        ws.row_dimensions[r_idx].height = 20

    # Legend
    legend_row = len(rows) + 5
    ws.cell(row=legend_row, column=1, value="Legend:").font = cell_font(bold=True, color="9D9D9D")
    legends = [
        (C_BEST_MAP,  "Best mAP@50"),
        (C_BEST_FPS,  "Best FPS"),
        (C_YOLO_BG,   "YOLO family"),
        (C_MOBILE_BG, "MobileNet family"),
    ]
    for i, (color, label) in enumerate(legends, 2):
        c = ws.cell(row=legend_row, column=i * 2 - 1)
        c.fill = fill(color)
        c.value = label
        c.font  = cell_font(color=C_TEXT, bold=False)
        c.alignment = center()
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(i * 2 - 1)].width = 18

    return len(rows)


# ── Sheet 2: Speed vs Accuracy Scatter ──────────────────────────────── #
def build_speed_accuracy_sheet(ws, headers, rows):
    ws.title = "Speed vs Accuracy"
    ws.sheet_view.showGridLines = False
    ws.tab_color = "4EC94E"

    eval_mode = is_eval_mode(headers)

    # Write data table for chart
    ws["A1"] = "Model"
    ws["B1"] = "FPS"
    ws["C1"] = "mAP@50 (%)" if eval_mode else "Avg dets"
    ws["D1"] = "Family"

    for cell in [ws["A1"], ws["B1"], ws["C1"], ws["D1"]]:
        cell.font = header_font()
        cell.fill = fill(C_HEADER_BG)
        cell.alignment = center()

    for r_idx, row in enumerate(rows, 2):
        fps = safe_float(row.get("FPS", ""))
        if eval_mode:
            acc = safe_float(row.get("mAP@50 (%)", ""))
        else:
            acc = safe_float(row.get("Avg dets", ""))

        ws.cell(row=r_idx, column=1, value=row.get("Model", "").strip())
        ws.cell(row=r_idx, column=2, value=fps)
        ws.cell(row=r_idx, column=3, value=acc)
        ws.cell(row=r_idx, column=4, value=row.get("Family", "").strip())

    # Scatter chart
    chart = ScatterChart()
    chart.title   = "Speed vs Accuracy  (top-right = best)"
    chart.style   = 10
    chart.x_axis.title = "FPS (higher = faster)"
    chart.y_axis.title = "mAP@50 (%)" if eval_mode else "Avg detections"
    chart.x_axis.numFmt = "0.0"
    chart.y_axis.numFmt = "0.0"
    chart.width  = 22
    chart.height = 16

    n = len(rows) + 1
    xvalues = Reference(ws, min_col=2, min_row=2, max_row=n)
    yvalues = Reference(ws, min_col=3, min_row=2, max_row=n)

    series = Series(yvalues, xvalues, title="Models")
    series.marker.symbol   = "circle"
    series.marker.size     = 8
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)

    ws.add_chart(chart, "F2")

    # Column widths
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12


# ── Sheet 3: Accuracy Comparison Bar Chart ──────────────────────────── #
def build_accuracy_sheet(ws, headers, rows):
    ws.title = "Accuracy Comparison"
    ws.sheet_view.showGridLines = False
    ws.tab_color = "9CDCFE"

    if not is_eval_mode(headers):
        ws["A1"] = "Accuracy metrics not available — run benchmark in Accuracy mode."
        ws["A1"].font = cell_font(color=C_ORANGE, bold=True)
        return

    # Data table
    cols = ["Model", "mAP@50 (%)", "mAP@50:95 (%)", "Precision (%)", "Recall (%)"]
    for c_idx, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font      = header_font()
        cell.fill      = fill(C_HEADER_BG)
        cell.alignment = center()

    for r_idx, row in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=row.get("Model", "").strip())
        for c_idx, col in enumerate(cols[1:], 2):
            ws.cell(row=r_idx, column=c_idx, value=safe_float(row.get(col, "")))

    n = len(rows) + 1

    # Grouped bar chart
    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Accuracy Metrics by Model"
    chart.style   = 10
    chart.y_axis.title  = "Score (%)"
    chart.x_axis.title  = "Model"
    chart.y_axis.numFmt = "0.0"
    chart.width  = 30
    chart.height = 16

    cats = Reference(ws, min_col=1, min_row=2, max_row=n)
    for c_idx, metric in enumerate(["mAP@50 (%)", "mAP@50:95 (%)", "Precision (%)", "Recall (%)"], 2):
        data = Reference(ws, min_col=c_idx, min_row=1, max_row=n)
        series = Series(data, title_from_data=True)
        chart.series.append(series)
    chart.set_categories(cats)
    chart.shape = 4

    ws.add_chart(chart, "G2")

    ws.column_dimensions["A"].width = 36
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 15


# ── Sheet 4: Speed Comparison ────────────────────────────────────────── #
def build_speed_sheet(ws, headers, rows):
    ws.title = "Speed Comparison"
    ws.sheet_view.showGridLines = False
    ws.tab_color = "CE9178"

    cols = ["Model", "FPS", "Avg ms", "Size (MB)"]
    for c_idx, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font      = header_font()
        cell.fill      = fill(C_HEADER_BG)
        cell.alignment = center()

    for r_idx, row in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=row.get("Model", "").strip())
        ws.cell(row=r_idx, column=2, value=safe_float(row.get("FPS", "")))
        ws.cell(row=r_idx, column=3, value=safe_float(row.get("Avg ms", "")))
        ws.cell(row=r_idx, column=4, value=safe_float(row.get("Size (MB)", "")))

    n = len(rows) + 1
    cats = Reference(ws, min_col=1, min_row=2, max_row=n)

    # FPS bar chart
    fps_chart = BarChart()
    fps_chart.type   = "col"
    fps_chart.title  = "Inference Speed — FPS by Model"
    fps_chart.style  = 10
    fps_chart.y_axis.title  = "Frames per Second"
    fps_chart.x_axis.title  = "Model"
    fps_chart.y_axis.numFmt = "0.0"
    fps_chart.width  = 28
    fps_chart.height = 14

    fps_data = Reference(ws, min_col=2, min_row=1, max_row=n)
    fps_series = Series(fps_data, title_from_data=True)
    fps_chart.series.append(fps_series)
    fps_chart.set_categories(cats)
    ws.add_chart(fps_chart, "F2")

    # Avg ms bar chart
    ms_chart = BarChart()
    ms_chart.type   = "col"
    ms_chart.title  = "Inference Latency — Avg ms by Model"
    ms_chart.style  = 10
    ms_chart.y_axis.title  = "Milliseconds per Image"
    ms_chart.x_axis.title  = "Model"
    ms_chart.y_axis.numFmt = "0.0"
    ms_chart.width  = 28
    ms_chart.height = 14

    ms_data = Reference(ws, min_col=3, min_row=1, max_row=n)
    ms_series = Series(ms_data, title_from_data=True)
    ms_chart.series.append(ms_series)
    ms_chart.set_categories(cats)
    ws.add_chart(ms_chart, "F20")

    ws.column_dimensions["A"].width = 36
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 12


# ── Sheet 5: Size vs Accuracy ────────────────────────────────────────── #
def build_size_accuracy_sheet(ws, headers, rows):
    ws.title = "Size vs Accuracy"
    ws.sheet_view.showGridLines = False
    ws.tab_color = "DCDCAA"

    eval_mode = is_eval_mode(headers)

    cols = ["Model", "Size (MB)", "mAP@50 (%)" if eval_mode else "FPS", "Family"]
    for c_idx, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font      = header_font()
        cell.fill      = fill(C_HEADER_BG)
        cell.alignment = center()

    for r_idx, row in enumerate(rows, 2):
        acc_col = "mAP@50 (%)" if eval_mode else "FPS"
        ws.cell(row=r_idx, column=1, value=row.get("Model", "").strip())
        ws.cell(row=r_idx, column=2, value=safe_float(row.get("Size (MB)", "")))
        ws.cell(row=r_idx, column=3, value=safe_float(row.get(acc_col, "")))
        ws.cell(row=r_idx, column=4, value=row.get("Family", "").strip())

    n = len(rows) + 1
    chart = ScatterChart()
    chart.title   = "Model Size vs " + ("mAP@50  (top-left = efficient)" if eval_mode else "FPS")
    chart.style   = 10
    chart.x_axis.title = "Model Size (MB)"
    chart.y_axis.title = "mAP@50 (%)" if eval_mode else "FPS"
    chart.x_axis.numFmt = "0.0"
    chart.y_axis.numFmt = "0.0"
    chart.width  = 22
    chart.height = 16

    xvalues = Reference(ws, min_col=2, min_row=2, max_row=n)
    yvalues = Reference(ws, min_col=3, min_row=2, max_row=n)
    series  = Series(yvalues, xvalues, title="Models")
    series.marker.symbol   = "diamond"
    series.marker.size     = 8
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    ws.add_chart(chart, "F2")

    ws.column_dimensions["A"].width = 36
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 14


# ── Sheet 6: Raw Data ────────────────────────────────────────────────── #
def build_raw_sheet(ws, headers, rows):
    ws.title = "Raw Data"
    ws.sheet_view.showGridLines = True
    ws.tab_color = "6D6D6D"

    for c_idx, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font      = Font(name="Segoe UI", bold=True, size=10)
        cell.fill      = fill("2D2D30")
        cell.alignment = center()
        ws.column_dimensions[get_column_letter(c_idx)].width = 18

    for r_idx, row in enumerate(rows, 2):
        for c_idx, col in enumerate(headers, 1):
            val = row.get(col, "").strip()
            f   = safe_float(val)
            cell = ws.cell(row=r_idx, column=c_idx, value=f if f is not None else val)
            cell.font = Font(name="Consolas", size=9)


# ── Main ─────────────────────────────────────────────────────────────── #
def build_workbook(csv_path: str, out_path: str):
    headers, rows = read_csv(csv_path)

    if not rows:
        print("ERROR: No data rows found in CSV.")
        sys.exit(1)

    eval_mode = is_eval_mode(headers)
    print(f"Loaded {len(rows)} model results  |  "
          f"mode={'accuracy+speed' if eval_mode else 'speed only'}")

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # Dark workbook theme
    ws_results = wb.create_sheet("Results")
    ws_speed_acc = wb.create_sheet("Speed vs Accuracy")
    ws_accuracy  = wb.create_sheet("Accuracy Comparison")
    ws_speed     = wb.create_sheet("Speed Comparison")
    ws_size_acc  = wb.create_sheet("Size vs Accuracy")
    ws_raw       = wb.create_sheet("Raw Data")

    build_results_sheet(ws_results, headers, rows, eval_mode)
    build_speed_accuracy_sheet(ws_speed_acc, headers, rows)
    build_accuracy_sheet(ws_accuracy, headers, rows)
    build_speed_sheet(ws_speed, headers, rows)
    build_size_accuracy_sheet(ws_size_acc, headers, rows)
    build_raw_sheet(ws_raw, headers, rows)

    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


def main():
    parser = argparse.ArgumentParser(description="Export benchmark CSV to Excel")
    parser.add_argument("--csv", required=True, help="Path to benchmark CSV file")
    parser.add_argument("--out", default=None,
                        help="Output .xlsx path (default: same location as CSV)")
    args = parser.parse_args()

    csv_path = Path(args.csv)
    if not csv_path.exists():
        print(f"ERROR: CSV not found: {csv_path}")
        sys.exit(1)

    if args.out:
        out_path = Path(args.out)
    else:
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = csv_path.parent / f"benchmark_{ts}.xlsx"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(str(csv_path), str(out_path))


if __name__ == "__main__":
    main()
